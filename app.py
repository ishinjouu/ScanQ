import streamlit as st
import pdfplumber
import pandas as pd
import numpy as np
import hashlib
import re
import requests
import json
import traceback
from difflib import get_close_matches
import pandas as pd
from typing import List
from io import BytesIO
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from difflib import SequenceMatcher

#---------------------- API --------------------------
from flask import Flask, request, jsonify

app = Flask(__name__)
#-----------------------------------------------------

def fix_split_standard_issue(row):
    item = str(row.get("Item", "")).strip()
    standard = str(row.get("Standard", "")).strip()

    # Case 1: Spacing single uppercase letter for Q & F at the start
    spacing = re.match(r"^([QF])\s+(.*)", item)
    if spacing:
        ch, rest = spacing.groups()
        item = rest
        standard = f"{ch} {standard}" if not standard.lower().startswith(ch.lower()) else standard

    # Case 2: Spacing single uppercase letter for Q & F at the end
    spacing = re.match(r"^(.*)\s+([QF])$", item)
    if spacing:
        rest, ch = spacing.groups()
        item = rest
        standard = f"{ch} {standard}" if not standard.lower().startswith(ch.lower()) else standard

    # Case 3: Trailing single uppercase letter in the middle
    middle = re.search(r"^(.*?)(?:\s)([A-Z])(?:\s)(.*)$", item)
    if middle:
        before, ch, after = middle.groups()
        item = (before + " " + after).strip()
        standard = f"{ch} {standard}".strip() if not standard.lower().startswith(ch.lower()) else standard

    # Case 4: Leading uppercase letter at the start
    leading = re.match(r"^([A-Z])\s+(.*)", item)
    if leading:
        ch, rest = leading.groups()
        item = rest
        standard = f"{ch}{standard}" if not standard.lower().startswith(ch.lower()) else standard

    # Case 5: Trailing single uppercase letter at the end
    trailing = re.match(r"^(.*)\s+([A-W])$", item)
    if trailing:
        rest, ch = trailing.groups()
        item = rest
        standard = f"{ch}{standard}" if not standard.lower().startswith(ch.lower()) else standard

    row["Item"] = item.strip()
    row["Standard"] = standard.strip()
    return row

# sanitazing json list from jenis_pengecekan
def sanitize_for_json(value):
    if isinstance(value, float) and (pd.isna(value) or not np.isfinite(value)):
        return None
    if isinstance(value, list):
        return [sanitize_for_json(v) for v in value]
    if isinstance(value, dict):
        return {k: sanitize_for_json(v) for k, v in value.items()}
    return value

# upload file in streamlite
def get_file_hash(file):
    return hashlib.md5(file.getvalue()).hexdigest()

# if string, then remove \n, '' & turn into lowercase
def normalize_text(text):
    if not isinstance(text, str):
        return text
    return text.replace('\n', '').strip().lower()

# reversing the "replacements"
def maybe_flip_text(text):
    if not isinstance(text, str) or not text.strip():
        return text
    raw = text.replace('\n', '').strip()
    reversed_text = raw[::-1]
    normalized_reversed = normalize_text(reversed_text)
    replacements = {
        "setup": "Set Up",
        "pu tes": "Set Up",
        "patrol": "Patrol",
        "1x/shift": "Patrol 1x/Shift",
        "job setup": "Job Set Up",
        "portal": "Patrol",
        "up": "Up",
        "1x/day": "Patrol 1x/Day",
        "shift": "Shift",
        "allpointifjobsetup": "All Point If Job Set Up",
        "4allpointifjobsetup": "All Point If Job Set Up"
    }
    if normalized_reversed in replacements:
        pretty_text = replacements[normalized_reversed]
        return pretty_text
    return text

# reversing a string / text
def reverse_text(text):
    if not isinstance(text, str):
        return text
    text = text.replace('\n', ' ')
    return text[::-1].strip()

# convert data into excel
def convert_df_to_excel(df):
    if df is None or df.empty:
        st.warning("⚠ Cleaned dataframe is empty or None. Skipping Excel export.")
        return None
    output = BytesIO()
    try:
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Sheet1')
        output.seek(0)
        wb = load_workbook(output)
        ws = wb.active
        for col_idx, column_cells in enumerate(ws.columns, 1):
            max_length = 0
            col_letter = get_column_letter(col_idx)
            col_name = column_cells[0].value
            for row_idx, cell in enumerate(column_cells, 1):
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                    if row_idx == 1:
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    else:
                        if col_name and str(col_name).strip().lower() == "section":
                            cell.alignment = Alignment(horizontal='left', vertical='center')
                        else:
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                except:
                    pass
                
            adjusted_width = max_length + 2
            ws.column_dimensions[col_letter].width = adjusted_width
        adjusted_output = BytesIO()
        wb.save(adjusted_output)
        adjusted_output.seek(0)
        return adjusted_output

    except Exception as e:
        st.error(f"💥 Excel writing failed: {e}")
        return None


# ---------- PDF Table Extraction ----------

def pisahkan_item_dan_extra(df):
    if "Item" in df.columns and "_2" not in df.columns:
        extracted = df["Item"].str.extract(r"^(.*?)(\s*\(E\d+\))$")
        df["Item"] = extracted[0].fillna(df["Item"]).str.strip()
        df["_2"] = extracted[1].fillna("").str.strip()
    return df

def extract_table_from_pdf(file):
    all_dataframes = []
    max_columns = 0


    with pdfplumber.open(file) as pdf:
        for page_num, page in enumerate(pdf.pages):
            tables = page.extract_tables()
            page_tables = []


            # st.write(f"📄 Halaman {page_num + 1}")


            # for table_idx, table in enumerate(tables):
            #     st.write(f"  ➤ Tabel {table_idx + 1} - Jumlah baris: {len(table)}")
            #     for i, row in enumerate(table):
            #         st.write(f"    Row {i}: {row}")


            for table_idx, table in enumerate(tables):
                if not table or len(table) < 2:
                    continue


                header_row_idx = None
                for idx, row in enumerate(table):
                    if row and any("Item" in str(cell) for cell in row):
                        header_row_idx = idx
                        break


                if header_row_idx is not None:
                    data = table[header_row_idx:]
                    expected_cols = 13
                    # header = [str(h).strip() if h is not None else "" for h in data[0]]
                    header = [deduplicate_words(str(h).strip()) if h is not None else "" for h in data[0]]


                    while len(header) < expected_cols:
                        header.append(f"Extra_{len(header)}")


                    normalized_data = []
                    for row in data[1:]:
                        padded_row = row + [""] * (len(header) - len(row))


                        target_idx = 2
                        if not padded_row[target_idx] or str(padded_row[target_idx]).strip() == "":
                            for val in padded_row[5:10]:  # kolom 6-10 tempat biasanya 'a' atau 'b'
                                # if isinstance(val, str) and val.strip().isalpha() and len(val.strip()) == 1:
                                val_clean = str(val).strip()
                                if val_clean.isdigit():
                                    padded_row[target_idx] = val_clean
                                    break

                        normalized_data.append(padded_row)


                    seen = {}
                    new_header = []
                    for col in header:
                        if col in seen:
                            seen[col] += 1
                            new_header.append(f"{col}_{seen[col]}")
                        else:
                            seen[col] = 0
                            new_header.append(col)


                    df = pd.DataFrame(normalized_data, columns=new_header)
                    page_tables.append(df)
                else:
                    df = pd.DataFrame(table)
                    page_tables.append(df)

                df["page_number"] = page_num + 1
                # page_tables.append(df)


            if page_tables:
                try:
                    merged_df = pd.concat(page_tables, axis=0, ignore_index=True)
                    all_dataframes.append(merged_df)
                except Exception as e:
                    st.warning(f"⚠️ Gagal merge tabel di halaman {page_num+1}: {e}")


    normalized_tables = []
    for df in all_dataframes:
        if df.shape[1] < max_columns:
            for i in range(df.shape[1], max_columns):
                df[f"Extra_{i}"] = ""
        normalized_tables.append(df)


    return pd.concat(normalized_tables, ignore_index=True) if normalized_tables else pd.DataFrame()


def merge_partial_rows(df, value_col="Standard", threshold=3):
    merged_rows = []
    buffer = None

    for _, row in df.iterrows():
        non_empty_cells = [str(v).strip() for v in row if str(v).strip().lower() not in ["", "none", "nan"]]
        
        if str(row.get("Item", "")).strip():
            if buffer is not None:
                merged_rows.append(buffer)
            buffer = row.copy()
        elif buffer is not None and len(non_empty_cells) <= threshold:
            # Baris lanjutan dengan isi sangat sedikit
            existing_val = str(buffer.get(value_col, "")).strip()
            combined_val = ", ".join(filter(None, [existing_val] + non_empty_cells))
            buffer[value_col] = combined_val.strip(", ")
        else:
            if buffer is not None:
                merged_rows.append(buffer)
                buffer = None
            merged_rows.append(row)  # anggap baris baru
    if buffer is not None:
        merged_rows.append(buffer)
    return pd.DataFrame(merged_rows)


def group_rows_by_item(df):
    if "Item" not in df.columns or "Standard" not in df.columns:
        return df
    grouped_rows = []
    buffer_row = None
    for idx, row in df.iterrows():
        item = str(row.get("Item", "")).strip()
        std = str(row.get("Standard", "")).strip()
        detail = str(row.get("Detail Standard", "")).strip() if "Detail Standard" in df.columns else ""
        if item != "":
            if buffer_row is not None:
                grouped_rows.append(buffer_row)
            buffer_row = row.copy()
        else:
            if buffer_row is not None:
                buffer_row["Standard"] = f"{buffer_row['Standard']}, {std}".strip(', ')
                if "Detail Standard" in buffer_row and detail:
                    buffer_row["Detail Standard"] = f"{buffer_row['Detail Standard']}, {detail}".strip(', ')
    if buffer_row is not None:
        grouped_rows.append(buffer_row)

    return pd.DataFrame(grouped_rows)

# ---------- Footer & Flip Cleaners ----------

def hapus_footer(df):
    keywords = ["keputusan", "keterangan", "approved", "checked", "disetujui", "diperiksa", "dibuat", "nama", "tanggal", "ttd"]
    
    cleaned_pages = []
    unique_pages = df["page_number"].unique()

    for page in unique_pages:
        page_df = df[df["page_number"] == page].copy() 
        footer_start_idx = None

        for idx in page_df.index:
            row = page_df.loc[idx]
            row_str = ' '.join([str(x).lower() for x in row if pd.notnull(x)])
            if any(k in row_str for k in keywords):
                footer_start_idx = idx
                break
        if footer_start_idx is not None:
            st.info(f"🧹 Footer detected at index {footer_start_idx}. Removing it.")
            page_df = page_df.loc[:footer_start_idx - 1].copy()
        else:
            st.info("✅ No footer found.")

        cleaned_pages.append(page_df)
    
    final_df = pd.concat(cleaned_pages, ignore_index=True)
    return final_df

def detect_and_fix_reversed_columns(df):
    replacements = {
        "setup": "Set Up",
        "pu tes": "Set Up",
        "patrol": "Patrol",
        "1x/shift": "Patrol 1x/Shift",
        "job setup": "Job Set Up",
        "portal": "Patrol",
        "up": "Up",
        "1x/day": "Patrol 1x/Day",
        "shift": "Shift",
        "allpointifjobsetup": "All Point If Job Set Up",
        "4allpointifjobsetup": "All Point If Job Set Up"
    }

    def try_fix_header(col_name):
        col_name = str(col_name)
        norm = normalize_text(col_name)
        rev = norm[::-1]
        if rev in replacements:
            return replacements[rev]
        elif norm in replacements:
            return replacements[norm]
        else:
            return col_name.strip()
        
    new_columns = [try_fix_header(col) for col in df.columns]
    df.columns = new_columns
    return df

def fill_patrol_column(df):
    df["_is_section"] = df["No."].astype(str).str.match(r'^\s*(I|II|III|IV|V|VI|VII|VIII|IX|X)\b')
    last_valid_patrol = None
    patrol_filled = []

    for is_section, patrol in zip(df["_is_section"], df["Patrol"]):
        if is_section:
            patrol_filled.append(patrol)
            last_valid_patrol = None
        else:
            if pd.notna(patrol):
                last_valid_patrol = patrol
                patrol_filled.append(patrol)
            else:
                patrol_filled.append(last_valid_patrol)

    df["Patrol"] = patrol_filled
    df.drop(columns=["_is_section"], inplace=True)
    return df

def fill_setup_from_patrol(df):
    if "Set Up" not in df.columns:
        df["Set Up"] = None
    if "Patrol" not in df.columns:
        st.warning("🛑 Kolom 'Patrol' tidak ditemukan.")
        return df
    if "No." not in df.columns:
        st.warning("🛑 Kolom 'No.' tidak ditemukan.")
        return df

    df["_is_section"] = df["No."].astype(str).str.match(r'^\s*(I|II|III|IV|V|VI|VII|VIII|IX|X)\b')
    df["_filled_patrol"] = df["Patrol"].ffill()

    setup_filled = []
    for is_section, patrol_val in zip(df["_is_section"], df["_filled_patrol"]):
        if is_section:
            setup_filled.append(None)
        else:
            if pd.notna(patrol_val) and str(patrol_val).strip() != "":
                setup_filled.append("Job Setup")
            else:
                setup_filled.append(None)

    df["Set Up"] = setup_filled
    df.drop(columns=["_is_section", "_filled_patrol"], inplace=True)
    return df

# ---------- Final Cleaning ----------

def hapus_ok_ng(text):
    if not isinstance(text, str):
        return text
    return re.sub(r'\bOK\b\s*[/-]?\s*\bN[G]?\b', '', text, flags=re.IGNORECASE).strip()

def fill_standard_from_job_setup(df):
    df = df.copy()
    if "Set Up" not in df.columns or "Standard" not in df.columns:
        return df
    last_valid_std = None
    for idx, row in df.iterrows():
        jenis = str(row.get("Set Up", "")).strip().lower()
        std = str(row.get("Standard", "")).strip()
        if jenis == "job setup":
            if std not in ["", "-", "nan", "none", None]:
                last_valid_std = std
            elif last_valid_std:
                df.at[idx, "Standard"] = last_valid_std
    return df

def normalisasi_m_notasi(standard):
    if not isinstance(standard, str):
        return standard
    pattern = r'\b[Mm]\s?-?\s?(\d{1,2})(?![\d.±°])\b'
    match = re.search(pattern, standard)
    if match:
        angka = match.group(1)
        standard = re.sub(pattern, f'[M{angka}]', standard)
    return standard

def deduplicate_words(text):
    if not isinstance(text, str):
        return text
    words = text.strip().split()
    deduped = []
    for word in words:
        if not deduped or deduped[-1] != word:
            deduped.append(word)
    return " ".join(deduped)

ROMAWI_PATTERN = re.compile(r"^(?=[MDCLXVI])(M{0,4}(CM|CD|D?C{0,3})(XC|XL|L?X{0,3})(IX|IV|V?I{0,3}))$")

def is_romawi(val):
    if isinstance(val, str):
        return ROMAWI_PATTERN.fullmatch(val.strip().upper()) is not None
    return False

def fill_down_except_romawi(df, kolom_target):
    df = df.copy()
    last_valid = None
    for idx, val in df[kolom_target].items():
        if pd.notnull(val) and str(val).strip() != "" and not is_romawi(val):
            last_valid = val
        elif pd.isnull(val) or str(val).strip() == "":
            df.at[idx, kolom_target] = last_valid
    return df

def gabungkan_kolom_item(df, kolom_Item='Item'):
    if kolom_Item not in df.columns:
        st.warning(f"🛑 Kolom '{kolom_Item}' tidak ditemukan.")
        return df

    idx_Item = df.columns.get_loc(kolom_Item)

    if idx_Item + 1 >= len(df.columns):
        st.warning("🛑 Tidak ada kolom setelah kolom Item.")
        return df

    kolom_1 = df.columns[idx_Item + 1]
    kolom_2 = df.columns[idx_Item + 2] if idx_Item + 2 < len(df.columns) else None

    gabung_kolom_1 = pd.isna(kolom_1) or str(kolom_1).strip() == ''
    gabung_kolom_2 = kolom_2 == '_1'

    if gabung_kolom_1:
        if gabung_kolom_2 and kolom_2 in df.columns:
            st.info(f"🔧 Gabungkan isi dari dua kolom setelah '{kolom_Item}' → format: 'Item (Tambahan1 - Tambahan2)'")

            def gabung(val_Item, val1, val2):
                vals = [str(v).strip() for v in [val1, val2] if pd.notna(v) and str(v).strip()]
                if vals:
                    return f"{val_Item} ({' - '.join(vals)})"
                return val_Item

            df[kolom_Item] = df.apply(lambda row: gabung(row[kolom_Item], row[kolom_1], row[kolom_2]), axis=1)
            df.drop(columns=[kolom_1, kolom_2], inplace=True)
        else:
            st.info(f"🔧 Gabungkan isi dari kolom tak bernama setelah '{kolom_Item}' → format: 'Item (Tambahan)'")

            def gabung(val_Item, val_samping):
                if pd.notna(val_samping) and str(val_samping).strip() != '':
                    return f"{val_Item} ({val_samping})"
                return val_Item

            df[kolom_Item] = df.apply(lambda row: gabung(row[kolom_Item], row[kolom_1]), axis=1)
            df.drop(columns=[kolom_1], inplace=True)

        # Rename any remaining unnamed columns
        df.columns = [f"col_{i}" if not c or pd.isna(c) else c for i, c in enumerate(df.columns)]
    else:
        st.info("ℹ️ Kolom setelah 'Item' punya nama, tidak digabung.")

    return df

def fill_item(df, kolom_item='Item', kolom_standard='Standard'):
    df = df.reset_index(drop=True)  
    def is_section(text):
        if isinstance(text, str):
            return bool(re.match(r'^[IVXLCDM]+\.', text.strip()))  # regex untuk format Romawi titik (mis. "I.", "II.", dll)
        return False
    for i in range(1, len(df)):
        item = df.at[i, kolom_item]
        std = df.at[i, kolom_standard] if kolom_standard in df.columns else ''
        if (pd.isna(item) or str(item).strip() == '') and pd.notna(std) and str(std).strip() != '':
            item_sebelumnya = df.at[i - 1, kolom_item]
            if not is_section(item_sebelumnya):
                df.at[i, kolom_item] = item_sebelumnya
    return df

def normalisasi_patrol(patrol_input):
    if not isinstance(patrol_input, str):
        return ""
    teks = patrol_input.lower().replace("\n", " ").replace("\r", " ").strip()
    mapping = {
        "Patrol 1x/Shift": [
            "1x/shift", "shift/1x", "1 shift", "x1/shift", "1x shift", "x1 / shift", "shift x1",
            "tfihs / x1", "x1 / tfihs", "tfihs/1x", "1x / tfihs", "tfihs x1",
            "tfihs / x1 tfihs / x1", "tfihs / x1 tfihs / x1 tfihs / x1", "1x / shift"
        ],
        "Patrol 1x/Day": [
            "1x/day", "day/1x", "1 day", "1x per day", "per day", "yad/x1", "x1/yad"
        ]
    }
    hasil = set()
    for kategori, variasi_list in mapping.items():
        for variasi in variasi_list:
            if variasi in teks:
                hasil.add(kategori)
    return ", ".join(sorted(hasil))

# label bolong b. dll
def isi_label_abjad_di_antara(df, kolom='Item', No='No.'):
    pola = re.compile(r'^([a-zA-Z])\.\s*(.+)$')
    new_items = []
    last_label = None
    for i in range(len(df)):
        if pd.notna(df.at[i, No]):
            last_label = None

        item = str(df.at[i, kolom]).strip()
        match = pola.match(item)
        if match:
            last_label = match.group(1)
            new_items.append(item)
        else:
            if last_label:
                new_items.append(f"{last_label}. {item}")
            else:
                new_items.append(item)
    df[kolom] = new_items
    return df

def gabungkan_kolom_mirip(df, target_col, alias_list):
    """
    Menggabungkan beberapa kolom dengan nama typo/mirip ke dalam satu kolom resmi (target_col).
    Nilai yang tidak kosong akan diprioritaskan dari kanan ke kiri (yang paling kanan akan menimpa).
    """
    if target_col not in df.columns:
        df[target_col] = None
    for col in alias_list:
        if col in df.columns:
            df[target_col] = df[target_col].combine_first(df[col])
            df.drop(columns=[col], inplace=True)
    return df

# -------------- Bersihkan Data Frame ---------------
def bersihkan_dataframe(df):
    df = detect_and_fix_reversed_columns(df)
    # -------------- Patrol Rusak --------------
    def perbaiki_patrol_mentah(teks):
        if not isinstance(teks, str):
            return teks

        # Normalisasi dasar
        teks = teks.lower()
        teks = re.sub(r"\s+", " ", teks).strip()
        teks = teks.replace("/", " ")  # ubah '/' jadi spasi biar regex gampang

        shift_synonyms = ["shift", "tfihs"]
        day_synonyms = ["day", "yad"]

        # Cari kombinasi x1 / x2 / x3 DAN kata shift/day (dalam bentuk apapun)
        found_x = re.search(r"x\d+", teks)
        if found_x:
            for syn in shift_synonyms:
                if syn in teks:
                    return "Patrol 1x/Shift"
            for syn in day_synonyms:
                if syn in teks:
                    return "Patrol 1x/Day"

        return teks
    try:
        if "No." in df.columns:
            df["No."] = df["No."].astype(str).str.replace(r'^(\d+)\s*(\w*)\.*', r'\1\2', regex=True)
        else:
            st.warning("🛑 Column 'No.' not found.")
    except Exception as e:
        st.warning(f"Cleaning 'No.' column failed: {e}")

    df.dropna(how='all', inplace=True)
    df = fill_setup_from_patrol(df)
    df = hapus_footer(df)
    df = merge_partial_rows(df)
    df = group_rows_by_item(df)
    df = pisahkan_item_dan_extra(df)
    if "Patrol" in df.columns:
        df["Patrol"] = df["Patrol"].apply(perbaiki_patrol_mentah)
        if "_is_section" in df.columns or "No." in df.columns:
            df = fill_patrol_column(df)
    else:
        df["Patrol"] = "Patrol 1x/Shift"

    for col in ["no_clean", "item_clean"]:
        if col in df.columns:
            df.drop(columns=[col], inplace=True)
    # cavity_columns = [col for col in df.columns if col.lower().startswith("cavity sample")]
    cavity_columns = [col for col in df.columns if any(key in col.lower() for key in ("cavity", "sample"))]
    if cavity_columns:
        df_cols = df.columns.tolist()
        first_cavity = cavity_columns[0]
        cavity_idx = df_cols.index(first_cavity)
        keep_cols = df_cols[:cavity_idx + 1]
        df = df[keep_cols]
        if len(cavity_columns) > 1:
            df.drop(columns=[col for col in cavity_columns[1:]], inplace=True)

    if "Standard" in df.columns:
        df["Standard"] = df["Standard"].apply(normalisasi_m_notasi)

        valid_patterns = re.compile(
            r'^(?:\s*'
            r'((?:min|max|\d+|[a-zA-Z]+)'
            r'(?:\s*[-–—]\s*(?:min|max|\d+|[a-zA-Z]+))?'
            r'(?:\s*,\s*(?:min|max|\d+|[a-zA-Z]+))*'
            r')\s*)$',
            re.IGNORECASE
        )

        def validate_standard(value):
            if isinstance(value, str):
                value_clean = value.strip()
                if valid_patterns.match(value_clean):
                    return value
                else:
                    return value  
            return value
        df["Standard"] = df["Standard"].apply(validate_standard)
    df = df.applymap(lambda x: '' if str(x).strip().lower() == 'none' else maybe_flip_text(x))

    kolom_bersihkan = ["Standard", "_1", "_2"]
    for kolom in kolom_bersihkan:
        if kolom in df.columns:
            df[kolom] = df[kolom].apply(hapus_ok_ng)
    df = fill_standard_from_job_setup(df)

    for kolom in ["No.","Standard", "item_check", "Jenis Point", "Alat Ukur"]:
        if kolom in df.columns:
            df = fill_down_except_romawi(df, kolom)
            df = df.apply(fix_split_standard_issue, axis=1)
    df = fill_standard_from_job_setup(df)
    df = fill_item(df, kolom_item='Item', kolom_standard='Standard')
    df = gabungkan_kolom_item(df, kolom_Item="Item")
    df = isi_label_abjad_di_antara(df, kolom='Item')
    if "patrol" in df.columns:
        df["patrol"] = df["patrol"].apply(lambda x: x.split() if isinstance(x, str) else x)
        df["patrol"] = df["patrol"].apply(normalisasi_patrol)
    # if = kolomnya ada 2
    df.columns = [col.strip().lower().replace('\n', ' ') for col in df.columns]
    df = gabungkan_kolom_mirip(df, "control_method", ["control method", "contorl method", "contro metho"])
    df.columns = [col.replace('_', ' ').title() for col in df.columns]
    # --------- Backup: Setup dan Patrol Default ---------
    if "Set Up" not in df.columns:
        df["Set Up"] = "Job Setup"

    if "Patrol" not in df.columns:
        df["Patrol"] = "Patrol 1x/Shift"

    return df

# ---------- Transform ke Format Final ----------

SECTION_REGEX = r'^\s*(I{1,3}|IV|V|VI{0,3}|VII{0,3}|VIII|IX|X)\s*[\.\-–]\s+.+'

def is_section_row(row):
    first_col = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
    print(f"[DEBUG] Cek kolom 1 section: {first_col}")
    return bool(re.match(SECTION_REGEX, first_col, re.IGNORECASE))

def extract_section_title(row):
    first_col = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
    return first_col

def clean_empty_rows(df_final, max_null=5):
    def is_row_useless(row):
        item_check = row.get("item_check", "")
        item_check_empty = pd.isna(item_check) or item_check.strip() == ""
        empty_count = row.isna().sum() + (row == '').sum()
        return item_check_empty and empty_count >= max_null
    df_cleaned = df_final[~df_final.apply(is_row_useless, axis=1)].copy()
    return df_cleaned

def merge_point_item(df):
    new_point_check = []
    new_item_check = []
    for point, item in zip(df["point_check"], df["item_check"]):
        item_str = str(item).strip()
        if re.match(r'^[a-zA-Z]\.', item_str):
            letter = item_str[0]
            remaining_item = re.sub(r'^[a-zA-Z]\.\s*', '', item_str, count=1)  
            new_point_check.append(f"{point}{letter}")
            new_item_check.append(remaining_item)
        else:
            new_point_check.append(point)
            new_item_check.append(item_str)
    df["point_check"] = new_point_check
    df["item_check"] = new_item_check
    return df

def roman_to_int(roman):
    roman = roman.upper()
    roman_dict = {'I': 1, 'V': 5, 'X': 10, 'L': 50, 'C': 100}
    result, prev = 0, 0
    for char in reversed(roman):
        val = roman_dict.get(char, 0)
        result += val if val >= prev else -val
        prev = val
    return result

def move_single_caps_to_note(row):
    standard = str(row.get("standard", "")).strip()
    item_check = str(row.get("item_check", "")).strip()
    catatan = str(row.get("catatan", "")).strip()

    tags_found = set()

    # Cek huruf kapital tunggal di kolom standard
    matches_standard = re.findall(r'(^|\s)([A-Z])(?![\w.])($|\s)', standard)
    for match in matches_standard:
        single_cap = match[1]
        if single_cap in ["F", "M", "Q"]:
            tags_found.add(f"[{single_cap}]")
    # Hapus huruf kapital dari kolom standard
    standard = re.sub(r'(^|\s)([A-Z])(?![\w.])($|\s)', ' ', standard).strip()

    # Cek huruf kapital tunggal di item_check dalam konteks akhir kalimat / spasi
    matches_item = re.findall(r'\b([A-Z])\b', item_check)
    for single_cap in matches_item:
        if single_cap in ["F", "M", "Q"]:
            tags_found.add(f"[{single_cap}]")

    # Tambahkan tag ke catatan (hindari duplikat)
    for tag in sorted(tags_found):
        if tag not in catatan:
            catatan = f"{catatan} {tag}".strip()

    row["standard"] = standard
    row["catatan"] = catatan
    return row

def copy_special_measurements_to_note(row):
    standard = str(row.get("standard", "")).strip()
    catatan = str(row.get("catatan", "")).strip()
    jenis_point = str(row.get("jenis_point", "")).strip()
    if jenis_point not in ["Dengan Ukur", "Dengan CMM"]:
        return row

    min_max_match = re.search(r"\b([Mm]in|[Mm]ax)\s*\d+(?:\.\d+)?", standard)
    if min_max_match:
        tag = f"[{min_max_match.group(0).strip()}]"
        if tag not in catatan:
            catatan += f" {tag}"

    size_patterns = [
        r"[Ø°]\d+(?:\.\d+)?\s*±\s*[+−-]?\d+(?:\.\d+)?",                                     # Ø10 ±0.1 atau °10 ± 0.5
        r"[Ø°]\d+(?:\.\d+)?\s*\(\s*\d+(?:\.\d+)?\s*~\s*[+−-]?\d+(?:\.\d+)?\s*\)",           # Ø6.1 ( 0 ~ +0.1 )
        r"[Ø°]\d+(?:\.\d+)?\s*\(\s*[+−-]?\d+(?:\.\d+)?\s*~\s*[+−-]?\d+(?:\.\d+)?\s*\)",     # Ø12.15 (-0.15 ~ +0.25)
        r"\d+(?:\.\d+)?º\s*±\s*\d+(?:\.\d+)?º",                                             # 15º ± 3º
    ]

    ukuran_found = None
    for pattern in size_patterns:
        match = re.search(pattern, standard)
        if match:
            ukuran_found = match.group(0).strip().replace("[", "").replace("]", "")
            break  

    if ukuran_found and ukuran_found not in catatan:
        catatan += f" {ukuran_found}"

    row["catatan"] = catatan.strip()
    return row

def normalize_note_tags(catatan):
    if not isinstance(catatan, str):
        return catatan
    all_tags = re.findall(r"\[([^\[\]]+)\]", catatan)
    allowed_tags = []
    for tag in all_tags:
        tag_clean = tag.strip()
        if tag_clean in ["F", "M", "Q"]:
            allowed_tags.append(f"[{tag_clean}]")
        elif re.match(r"(?i)^min\s+\d+(\.\d+)?$", tag_clean):
            allowed_tags.append(f"[{tag_clean}]")
        elif re.match(r"(?i)^max\s+\d+(\.\d+)?$", tag_clean):
            allowed_tags.append(f"[{tag_clean}]")

    for tag in all_tags:
        catatan = catatan.replace(f"[{tag}]", "")

    ordered = []
    for tag in ["[F]", "[M]", "[Q]"]:
        if tag in allowed_tags:
            ordered.append(tag)
    for tag in allowed_tags:
        if tag not in ["[F]", "[M]", "[Q]"]:
            ordered.append(tag)

    catatan_clean = catatan.strip()
    result = " ".join(ordered)
    if catatan_clean:
        result = f"{result} {catatan_clean}"

    return result.strip()

def parse_standard_value(row):
    standard = str(row.get("standard", "")).strip()
    jenis_point = str(row.get("jenis_point", "")).strip().lower()

    # Ganti koma dengan titik jika jenis_point adalah "dengan ukur" atau "dengan cmm"
    if jenis_point in ["dengan ukur", "dengan cmm"]:
        standard = standard.replace(",", ".")

    if jenis_point == "Tanpa Ukur":
        return pd.Series([None, None, None], index=["std_value", "std_min", "std_max"])

    # 1. 0 ±0.2
    match1 = re.search(r'(\d+(?:\.\d+)?)\s*±\s*(\d+(?:\.\d+)?)', standard)
    if match1:
        nominal = float(match1.group(1))
        delta = float(match1.group(2))
        return pd.Series([nominal, -delta, delta], index=["std_value", "std_min", "std_max"])

    # 2. Ø10 ±0.1 atau °5.5 ±0.2
    match2 = re.match(r'^[Ø°]\s*(\d+(?:\.\d+)?)\s*±\s*([+−-]?\d+(?:\.\d+)?)$', standard)
    if match2:
        nominal = float(match2.group(1))
        delta = float(match2.group(2).replace("−", "-"))
        return pd.Series([nominal, -abs(delta), abs(delta)], index=["std_value", "std_min", "std_max"])

    # 3. Ø6.1 ( 0 ~ +0.1 )
    match3 = re.match(r'^[Ø°]?\s*(-?\d+(?:\.\d+)?)\s*\(\s*([+-]?\d+(?:\.\d+)?)\s*~\s*([+-]?\d+(?:\.\d+)?)\s*\)', standard)
    if match3:
        nominal = float(match3.group(1))
        lower = float(match3.group(2))
        upper = float(match3.group(3))
        return pd.Series([nominal, lower, upper], index=["std_value", "std_min", "std_max"])

    # 4. [ 163.89 ± 0.25 ]
    match4 = re.match(r'^\[\s*(-?\d+(?:\.\d+)?)\s*±\s*([\d.]+)\s*\]$', standard)
    if match4:
        nominal = float(match4.group(1))
        delta = float(match4.group(2))
        return pd.Series([nominal, -delta, delta], index=["std_value", "std_min", "std_max"])

    # 5. Min/Max
    match5 = re.search(r'\b(Min|Max)\s*(\d+(?:\.\d+)?)\b', standard, re.IGNORECASE)
    if match5:
        kind = match5.group(1).lower()
        value = float(match5.group(2))
        return pd.Series([0, value if kind == "min" else 0, value if kind == "max" else 0], index=["std_value", "std_min", "std_max"])

    # 6. ( Reff : 0 ~ +0.5 )
    match6 = re.match(r'^\(.*?:\s*([+-]?\d+(?:\.\d+)?)\s*~\s*([+-]?\d+(?:\.\d+)?)\s*\)$', standard)
    if match6:
        lower = float(match6.group(1))
        upper = float(match6.group(2))
        return pd.Series([0, lower, upper], index=["std_value", "std_min", "std_max"])

    # 7. [0 (0 ~ +0.3]
    match7 = re.match(r'^\[\s*(-?\d+(?:\.\d+)?)\s*\(\s*([+-]?\d+(?:\.\d+)?)\s*~\s*([+-]?\d+(?:\.\d+)?).*$', standard)
    if match7:
        nominal = float(match7.group(1))
        lower = float(match7.group(2))
        upper = float(match7.group(3))
        return pd.Series([nominal, lower, upper], index=["std_value", "std_min", "std_max"])

    # 8. [Ø30.5 ± 0.2]
    match8 = re.match(r'^\[\s*[Ø°]?\s*(\d+(?:\.\d+)?)\s*±\s*([+−-]?\d+(?:\.\d+)?)\s*\]$', standard)
    if match8:
        nominal = float(match8.group(1))
        delta = float(match8.group(2).replace("−", "-"))
        return pd.Series([nominal, -abs(delta), abs(delta)], index=["std_value", "std_min", "std_max"])

    # 9. [reff. 0 (0 ~ +0.3)]
    match9 = re.match(r'^\[\s*.*?(-?\d+(?:\.\d+)?)\s*\(\s*([+-]?\d+(?:\.\d+)?)\s*~\s*([+-]?\d+(?:\.\d+)?)\s*\)\s*\]$', standard)
    if match9:
        nominal = float(match9.group(1))
        lower = float(match9.group(2))
        upper = float(match9.group(3))
        return pd.Series([nominal, lower, upper], index=["std_value", "std_min", "std_max"])

    # 10. ( Reff. 0 (0 ~ +0.3))
    match10 = re.match(r'^\(\s*.*?(-?\d+(?:\.\d+)?)\s*\(\s*([+-]?\d+(?:\.\d+)?)\s*~\s*([+-]?\d+(?:\.\d+)?)\s*\)\s*\)$', standard)
    if match10:
        nominal = float(match10.group(1))
        lower = float(match10.group(2))
        upper = float(match10.group(3))
        return pd.Series([nominal, lower, upper], index=["std_value", "std_min", "std_max"])

    # 11. 1 [0 ~ +0.5]
    match11 = re.match(r'^(-?\d+(?:\.\d+)?)\s*\[\s*([+-]?\d+(?:\.\d+)?)\s*~\s*([+-]?\d+(?:\.\d+)?)\s*\]$', standard)
    if match11:
        nominal = float(match11.group(1))
        lower = float(match11.group(2))
        upper = float(match11.group(3))
        return pd.Series([nominal, lower, upper], index=["std_value", "std_min", "std_max"])

    # 12. Ø7 [-0.3 ~ 0]
    match12 = re.match(r'^[Ø°]?\s*(-?\d+(?:\.\d+)?)\s*\[\s*([+-]?\d+(?:\.\d+)?)\s*~\s*([+-]?\d+(?:\.\d+)?)\s*\]$', standard)
    if match12:
        nominal = float(match12.group(1))
        lower = float(match12.group(2))
        upper = float(match12.group(3))
        return pd.Series([nominal, lower, upper], index=["std_value", "std_min", "std_max"])

    # 13. [Ø5.5 [-0.3 ~ 0]
    match13 = re.match(r'^\[\s*[Ø°]?\s*(\d+(?:\.\d+)?)\s*\[\s*([+-]?\d+(?:\.\d+)?)\s*~\s*([+-]?\d+(?:\.\d+)?).*$', standard)
    if match13:
        nominal = float(match13.group(1))
        lower = float(match13.group(2))
        upper = float(match13.group(3))
        return pd.Series([nominal, lower, upper], index=["std_value", "std_min", "std_max"])
   
    # 14. Reff. 0 ( 0 ~ +0.5 )
    match14 = re.match(r'^.*?(-?\d+(?:\.\d+)?)\s*\(\s*([+-]?\d+(?:\.\d+)?)\s*~\s*([+-]?\d+(?:\.\d+)?)\s*\)$',standard)
    if match14:
        nominal = float(match14.group(1))
        lower = float(match14.group(2))
        upper = float(match14.group(3))
        return pd.Series([nominal, lower, upper], index=["std_value", "std_min", "std_max"])
    
    # 15. Format: 15º ± 3º
    match15 = re.match(r'^(\d+(?:\.\d+)?)\s*º?\s*[±+]\s*(\d+(?:\.\d+)?)\s*º?$', standard)
    if match15:
        value = float(match15.group(1))
        margin = float(match15.group(2))
        return pd.Series([value, -margin, margin], index=["std_value", "std_min", "std_max"])

    # 16. Ambil angka dari bagian akhir teks seperti "Tidak ambles / minus 0 ( 0 ~ +0.5 )"
    match16 = re.search(r'(\d+(?:\.\d+)?)\s*\(\s*([+-]?\d+(?:\.\d+)?)\s*~\s*([+-]?\d+(?:\.\d+)?)\s*\)', standard)
    if match16:
        nominal = float(match16.group(1))
        lower = float(match16.group(2))
        upper = float(match16.group(3))
        return pd.Series([nominal, lower, upper], index=["std_value", "std_min", "std_max"])
    
    # 17. Format seperti "Reff. 0 ( -0.3 ~ 0 )" atau "Reff. 0 ( 0 ~ +0.5 )"
    match17 = re.search(
        r'reff\.*\s*([+-]?\d+(?:[.,]\d+)?)\s*\(\s*([+-]?\d+(?:[.,]\d+)?)\s*~\s*([+-]?\d+(?:[.,]\d+)?)\s*\)',
        standard,
        re.IGNORECASE
    )
    if match17:
        nominal = float(match17.group(1).replace(",", "."))
        lower = float(match17.group(2).replace(",", "."))
        upper = float(match17.group(3).replace(",", "."))
        return pd.Series([nominal, lower, upper], index=["std_value", "std_min", "std_max"])

    return pd.Series([None, None, None], index=["std_value", "std_min", "std_max"])

def fill_empty_catatan_from_group(df):
    df = df.copy()

    for idx, row in df.iterrows():
        if row.get("jenis_point") not in ["Dengan Ukur", "Dengan CMM"]:
            continue
        if row.get('catatan') and row['catatan'].strip() not in ["-", ""]:
            continue

        item = row['item_check']
        point_prefix = re.match(r'^(\d+[a-zA-Z]*)', str(row['point_check']))
        point_prefix = point_prefix.group(1) if point_prefix else ""

        for j, ref_row in df.iterrows():
            if j == idx:
                continue
            if ref_row.get("jenis_point") not in ["Dengan Ukur", "Dengan CMM"]:
                continue

            ref_prefix = re.match(r'^(\d+[a-zA-Z]*)', str(ref_row['point_check']))
            ref_prefix = ref_prefix.group(1) if ref_prefix else ""

            if (
                ref_prefix == point_prefix and
                isinstance(ref_row['item_check'], str)
            ):
                sim = SequenceMatcher(None, ref_row['item_check'], item).ratio()
                if sim >= 0.8 and ref_row['catatan']:
                    df.at[idx, 'catatan'] = ref_row['catatan']
                    break

    return df

# ----------- Validsi -----------
def find_mid_sequence_breaks(df):
    suspicious_indexes = []
    pattern = r"\(E(\d+)\)"
    grouped = {}
    for idx, row in df.iterrows():
        item = str(row.get("item_check", "")).strip()
        base = re.sub(pattern, "", item).strip().lower()
        match = re.search(pattern, item)
        if base not in grouped:
            grouped[base] = []
        grouped[base].append((idx, item, int(match.group(1)) if match else None))

    for base, items in grouped.items():
        items.sort(key=lambda x: x[0])
        numbers = [num for _, _, num in items if num is not None]
        if len(numbers) < 2:
            continue
        min_e = min(numbers)
        max_e = max(numbers)

        for idx, item_text, num in items:
            if num is None:
                pos = [i for i, (_, _, n) in enumerate(items) if n is not None]
                if not pos:
                    continue
                first = pos[0]
                last = pos[-1]
                current_pos = items.index((idx, item_text, num))
                if first < current_pos < last:
                    suspicious_indexes.append(idx)

    return suspicious_indexes
    # ----------- Validsi -----------

def transform_to_final_format(df):
    df.columns = [col.strip().replace('\n', ' ').title() for col in df.columns]
    if "Control Method" not in df.columns:
        df["Control Method"] = np.nan
    df = df.replace(to_replace=["", "nan", "None"], value=np.nan)
   
    def bersihkan_control_method_bocor(text):
        if not isinstance(text, str):
            return text
        # Hapus mulai dari OK sampai NG atau N
        cleaned = re.sub(r"\bOK\b.*?\bN[G]?\b", "", text, flags=re.IGNORECASE).strip()
        return cleaned

    df["Control Method"] = df["Control Method"].fillna(method="ffill")
    df["Control Method"] = df["Control Method"].apply(bersihkan_control_method_bocor)
    df = df.replace(to_replace=["", "nan", "None"], value=np.nan)

    dengan_ukur_keywords = ["caliper", "hg", "depth cal", "pitch dial", "rough. t", "hitung", "depth clp", "height g", "dial g"]
    tanpa_ukur_keywords = [
        "visual", "pg", "snap g.", "visual & punch", "visual + kikir", "visual & kikir",
        "machining test", "visual ( reff. master rough.)", "insp. jig", "finishing test"
    ]
    dengan_cmm_keywords = ["cmm"]

    final_rows = []
    current_section = ""
    last_valid_item = ""
    last_control_method = ""

    for idx, row in df.iterrows():
        if is_section_row(row):
            maybe_section = extract_section_title(row)
            print(f"🔍 [Row {idx}] Detected section → {maybe_section}")
            if maybe_section.strip() and len(maybe_section.strip()) > 5:
                current_section = maybe_section.strip()
            continue

        if row.isna().all():
            continue

        raw_point_check = str(row.get("No.", "")).strip()
        capital_letters = ''.join([c for c in raw_point_check if c.isalpha() and c.isupper()])
        cleaned_point_check = ''.join([c for c in raw_point_check if not (c.isalpha() and c.isupper())]).strip()
        extracted_note = f"[{capital_letters}]" if capital_letters else ""

        original_note = str(row.get("Note", "")).strip() if pd.notna(row.get("Note")) else ""
        final_note = f"{original_note} {extracted_note}".strip() if extracted_note else original_note

        item_raw = row.get("Item", np.nan)
        extra1 = row.get("_1", row.get("1", np.nan))
        extra2 = row.get("_2", row.get("2", np.nan))
        item = str(item_raw).strip() if pd.notna(item_raw) else ""
        extra1 = str(extra1).strip() if pd.notna(extra1) else ""
        extra2 = str(extra2).strip() if pd.notna(extra2) else ""
        if item:
            last_valid_item = item
        extras = []
        if extra1:
            extras.append(extra1)
        if extra2 and f"({extra2})" not in item:
            extras.append(extra2)

        item_check = f"{item or last_valid_item}"
        if extras:
            item_check += " (" + ", ".join(extras) + ")"

        control_method_raw = row.get("Control Method", "")
        control_method = str(control_method_raw).strip() if pd.notna(control_method_raw) else ""
        if not control_method:
            control_method = last_control_method
        control_method_lower = control_method.lower()

        if any(keyword in control_method_lower for keyword in dengan_ukur_keywords):
            jenis_point = "Dengan Ukur"
        elif any(keyword in control_method_lower for keyword in tanpa_ukur_keywords):
            jenis_point = "Tanpa Ukur"
        elif any(keyword in control_method_lower for keyword in dengan_cmm_keywords):
            jenis_point = "Dengan CMM"
        else:
            jenis_point = "Lainnya"

        # qtime_checked = st.session_state.get(f"qtime_{idx}", False)
        # check100_checked = st.session_state.get(f"check100_{idx}", False)

        jenis_pengecekan = [
            v.strip()
            for source in ["Set Up", "Patrol"]
            for v in str(row.get(source, "")).split(",")
            if v.strip() and v.strip() != "-"
        ]

        # Deteksi huruf kapital spesial di item_check (Qtime & 100%)
        if jenis_point == "Tanpa Ukur":
            tokens = item_check.split()
            if "T" in tokens:
                jenis_pengecekan.append("Q-time")
                tokens.remove("T")
            if "C" in tokens:
                jenis_pengecekan.append("100%")
                tokens.remove("C")
            item_check = " ".join(tokens)
        # if qtime_checked:
        #     jenis_pengecekan.append("Q-time")
        # if check100_checked:
        #     jenis_pengecekan.append("100%")
        jenis_pengecekan = list(dict.fromkeys(jenis_pengecekan))

        final_row = {
            "section": current_section,
            "point_check": cleaned_point_check,
            "jenis_point": jenis_point,
            "catatan": final_note,
            "item_check": item_check,
            "control_method": control_method or last_control_method,
            "standard": str(row.get("Standard", "")).strip(),
            "jenis_pengecekan": jenis_pengecekan
            # "Q-time": qtime_checked,
            # "check_100": check100_checked,
        }
        final_rows.append(final_row)
    df_result = pd.DataFrame(final_rows)
   
    def apply_note_transformations(row):
        if row["jenis_point"] in ["Dengan Ukur", "Dengan CMM"]:
            row = copy_special_measurements_to_note(row)
        row = move_single_caps_to_note(row)
        row["catatan"] = normalize_note_tags(row["catatan"])
        return row

    df_result = df_result.apply(apply_note_transformations, axis=1)

    map_jenis2 = {}
    if "Item" in df.columns and "Patrol" in df.columns:
        df["Patrol"] = df["Patrol"].fillna(method="ffill")
        for _, row2 in df.iterrows():
            item_key = str(row2.get("Item", "")).strip()
            extra = str(row2.get("_2", "")).strip()
            item_check_key = f"{item_key} ({extra})" if extra else item_key
            jenis2 = str(row2.get("Patrol", "")).strip()
            if item_check_key and jenis2:
                map_jenis2[item_check_key] = jenis2

    def update_jenis_pengecekan(row):
        current = row.get("jenis_pengecekan", [])
        if isinstance(current, str):
            current = [current] if current else []
        if not isinstance(current, list):
            current = list(current)
        tambahan = map_jenis2.get(row["item_check"])
        if tambahan and tambahan not in current:
            current.append(tambahan)
        return current

    df_result["jenis_pengecekan"] = df_result.apply(update_jenis_pengecekan, axis=1)

    def convert_roman_section_to_number(text):
        if pd.isna(text):
            return text
        match = re.match(r'^\s*([IVXLCDM]+)\s*\.\s+(.*)', str(text).strip(), re.IGNORECASE)
        if match:
            roman, title = match.groups()
            try:
                number = roman_to_int(roman)
                return f"{number}. {title.strip()}"
            except:
                return text
        return text

    df_result["section"] = df_result["section"].apply(convert_roman_section_to_number)

    def final_cleanup(df):
        def clean_catatan(row):
            catatan = str(row["catatan"]).strip()
            jenis = str(row["jenis_point"]).strip()
            has_tag = any(tag in catatan for tag in ["[F]", "[M]", "[Q]"])
            kosong = catatan == ""
            if jenis == "Tanpa Ukur":
                return "-" if kosong and not has_tag else catatan
            elif jenis in ["Dengan Ukur", "Dengan CMM"]:
                return "-" if kosong else catatan
            return catatan
        df["catatan"] = df.apply(clean_catatan, axis=1)
        cols_to_fill = [
            "section", "point_check", "jenis_point", "item_check",
            "control_method"
        ]
        for col in cols_to_fill:
            if col in df.columns:
                df[col] = df[col].replace(["", "nan", "None"], np.nan).fillna("-")
       
        if "jenis_pengecekan" in df.columns:
            df["jenis_pengecekan"] = df["jenis_pengecekan"].apply(
                lambda val: [v for v in val if v != "-"] if isinstance(val, list) else ["-"]
            )
            df["jenis_pengecekan"] = df["jenis_pengecekan"].apply(
                lambda val: val if val else ["-"]
            )

        return df

    def move_m_from_standard_to_note(row):
        std = str(row["standard"]).strip()
        catatan = str(row["catatan"]).strip()
        match = re.findall(r'\[M\d{1,2}\]', std)
        if match:
            for m_tag in match:
                std = std.replace(m_tag, "").strip()
                catatan = catatan.replace(m_tag, "").strip()
            catatan = f"{' '.join(match)} {catatan}".strip()
        row["standard"] = std
        row["catatan"] = catatan
        return row
   
    # --------------- Temukan duplikat sama persis -----------------
    def find_exact_duplicates(df):
        duplicate_indexes = []
        seen_keys = set()
        key_columns = [
            "section", "point_check", "jenis_point", "catatan",
            "item_check", "control_method", "std_value", "std_min", "std_max"
        ]
        for idx, row in df.iterrows():
            row_key = tuple(
                str(row.get(col, "")).strip().lower().replace(", ", ",").replace(" ,", ",").replace("\n", " ")
                if not pd.isna(row.get(col)) else ""
                for col in key_columns
            )
            if row_key in seen_keys:
                duplicate_indexes.append(idx)
            else:
                seen_keys.add(row_key)

        return duplicate_indexes

    def find_incomplete_duplicates(df):
        all_items = df["item_check"].tolist()
        suspicious_indexes = []

        for idx, row in df.iterrows():
            item_check = row["item_check"]
            jenis_pengecekan = row.get("jenis_pengecekan", [])

            if isinstance(jenis_pengecekan, str):
                jenis_pengecekan = [v.strip() for v in jenis_pengecekan.split(",") if v.strip()]
            elif not isinstance(jenis_pengecekan, list):
                jenis_pengecekan = [str(jenis_pengecekan).strip()]

            is_qtime_missing = "Q-time" not in jenis_pengecekan
            is_check100_missing = "100%" not in jenis_pengecekan

            tokens = item_check.strip().split()
            if "T" in tokens and is_qtime_missing:
                suspicious_indexes.append(idx)
                continue
            if "C" in tokens and is_check100_missing:
                suspicious_indexes.append(idx)
                continue

            is_suspect = (
                not any(char in item_check for char in "()") and
                get_close_matches(item_check, all_items, n=2, cutoff=0.85) and
                (
                    not jenis_pengecekan or
                    jenis_pengecekan == ["-"] or
                    any(pd.isna(j) or str(j).lower() == "nan" for j in jenis_pengecekan)
                )
            )

            if is_suspect:
                suspicious_indexes.append(idx)
        return suspicious_indexes
   
    # ---------------------- validasi ----------------------
    def find_invalid_format_rows(df):
        invalid_indexes = []

        for idx, row in df.iterrows():
            jenis_pengecekan = row.get("jenis_pengecekan", [])

            if isinstance(jenis_pengecekan, str):
                jenis_pengecekan = [v.strip() for v in jenis_pengecekan.split(",") if v.strip()]
            elif not isinstance(jenis_pengecekan, list):
                if pd.isna(jenis_pengecekan):
                    jenis_pengecekan = []
                else:
                    jenis_pengecekan = [str(jenis_pengecekan).strip()]

            # ❌ Kalau ADA "nan" di dalam list, maka baris invalid
            if any(str(j).strip().lower() == "nan" for j in jenis_pengecekan):
                invalid_indexes.append(idx)

        return invalid_indexes
   
    def find_suspicious_material_item(df):
        suspicious_indexes = []
        for idx, row in df.iterrows():
            item = str(row.get("item_check", "")).strip().lower()
            if "material" in item:
                tokens = item.split()
                if len(tokens) <= 3:
                    suspicious_indexes.append(idx)
        return suspicious_indexes
    # ---------------------- validasi ----------------------

    parsed_std = df_result.apply(parse_standard_value, axis=1)
    df_result[["std_value", "std_min", "std_max"]] = parsed_std[["std_value", "std_min", "std_max"]]
    df_result = fill_empty_catatan_from_group(df_result)
    df_result = df_result.apply(move_m_from_standard_to_note, axis=1)

    mask_diukur = df_result["jenis_point"].isin(["Dengan Ukur", "Dengan CMM"])
    df_result.loc[mask_diukur, "standard"] = None
    mask_tanpa_ukur = df_result["jenis_point"] == "Tanpa Ukur"
    parsed_std.loc[mask_tanpa_ukur, ["std_value", "std_min", "std_max"]] = [None, None, None]
    df_result[["std_value", "std_min", "std_max"]] = parsed_std[["std_value", "std_min", "std_max"]]

    df_result["jenis_point"] = df_result["jenis_point"].replace("Lainnya", np.nan).fillna(method="ffill")
    df_result["control_method"] = df_result["control_method"].replace(["", "nan", "None"], np.nan).fillna(method="ffill")
   
    df_result["standard_asli"] = df_result["standard"]
    mask_diukur = df_result["jenis_point"].isin(["Dengan Ukur", "Dengan CMM"])
    mask_tanpa_ukur = df_result["jenis_point"] == "Tanpa Ukur"
    df_result.loc[mask_tanpa_ukur, "standard"] = df_result.loc[mask_tanpa_ukur, "standard_asli"]
    df_result.drop(columns="standard_asli", inplace=True)
    df_result["point_check"] = df_result["point_check"].replace(["", " ", "nan", "None"], np.nan).fillna(method="ffill")

    df_result = clean_empty_rows(df_result)
    df_result = merge_point_item(df_result)
    df_result = final_cleanup(df_result)
 
    # 🚨 VALIDASI
    df_result["status"] = "valid"
    df_result.loc[find_exact_duplicates(df_result), "status"] = "duplikat"
    df_result.loc[find_incomplete_duplicates(df_result), "status"] = "duplikat_parsing"
    df_result.loc[find_invalid_format_rows(df_result), "status"] = "salah_format"
    df_result.loc[find_mid_sequence_breaks(df_result), "status"] = "salah_format"
    suspect_material = find_suspicious_material_item(df_result)
    df_result.loc[suspect_material, "status"] = "salah_format"
    # Hapus otomatis
    df_result = df_result[df_result["status"] == "valid"].reset_index(drop=True)

    return df_result

# ---------- Streamlit UI ----------

st.set_page_config(page_title="Check Sheet QFORM", layout="wide")
st.title("📄 DEBUG CHECK SHEET SCAN QFORM")

uploaded_file = st.file_uploader("📤 Upload PDF file", type="pdf")

if uploaded_file:
    file_hash = get_file_hash(uploaded_file)

    if "last_file_hash" not in st.session_state or st.session_state.last_file_hash != file_hash:
        st.session_state.last_file_hash = file_hash
        for key in ["df_final_data", "st.session_state.df_final_data", "show_updated_table"]:
            st.session_state.pop(key, None)
        st.cache_data.clear()

    try:
        df = extract_table_from_pdf(uploaded_file)
        if df.empty:
            st.warning("❌ No tables detected in the PDF.")
        else:
            df_cleaned = bersihkan_dataframe(df.copy())
            st.subheader("🚀 Cleaned Table")
            st.dataframe(df_cleaned, use_container_width=True, hide_index=True)

            if "df_final_data" not in st.session_state:
                st.session_state.df_final_data = transform_to_final_format(df_cleaned)
# ---------------------- validasi --------------------------------------------------------------------
            # 🚨 Validasi status data (duplikat vs valid)
            df_validasi = st.session_state.df_final_data
            total = len(df_validasi)
            duplikat_count = (df_validasi["status"] == "duplikat").sum()
            suspect_count = (df_validasi["status"] == "duplikat_parsing").sum()
            format_error_count = (df_validasi["status"] == "salah_format").sum()
            valid_count = total - duplikat_count - suspect_count - format_error_count

            st.info(f"""
            📊 *Validasi Baris*
            - Total: {total}
            - 🟢 Valid: {valid_count}
            - 🟡 Duplikat normal: {duplikat_count}
            - 🟠 Dugaan parsing rusak: {suspect_count}
            - ❌ Salah format: {format_error_count}
            """)

            if suspect_count > 0:
                st.warning("⚠️ Ditemukan baris yang kemungkinan hasil merge/parsing tidak sempurna.")

            df_validasi = st.session_state.df_final_data
            error_rows = df_validasi[df_validasi["status"].isin(["duplikat", "duplikat_parsing", "salah_format"])]
            error_count = len(error_rows)
            if error_count > 0:
                st.warning(f"⚠️ Ditemukan {error_count} baris tidak valid (duplikat atau salah format). Data akan diabaikan saat ekspor.")
# ---------------------- validasi ---------------------------------------------------------------------
            st.subheader("📊 Final Format")
            st.dataframe(
                st.session_state.df_final_data.reset_index(drop=True),
                use_container_width=True,
                hide_index=True
            )
            if st.button("♻️ Reset"):
                st.cache_data.clear()
                for key in ["df_final_data", "st.session_state.df_final_data", "show_updated_table", "last_file_hash"]:
                    st.session_state.pop(key, None)
                st.rerun()

    except Exception as e:
        st.error(f"🔥 Error while processing the file:\n\n{e}")
        st.text("📄 Traceback log:")
        st.text(traceback.format_exc())

# --- Endpoint API utama ----------------------------------------------------
@app.route("/api/proses_file", methods=["POST"])
def proses_file():
    if 'file' not in request.files:
        return jsonify({"error": "❌ Tidak ada file dikirim"}), 400

    file = request.files['file']
    try:
        df_raw = extract_table_from_pdf(file)
        df_clean = bersihkan_dataframe(df_raw)
        df_final = transform_to_final_format(df_clean)
        return jsonify(df_final.to_dict(orient="records"))
    except Exception as e:
        return jsonify({"error": f"🔥 Gagal proses file: {str(e)}"}), 500

# if __name__ == "__main__":
#     app.run(host="0.0.0.0", port=2051, debug=True)
